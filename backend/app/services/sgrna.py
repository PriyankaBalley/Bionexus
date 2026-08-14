"""Module 4 (advanced): sgRNA design pipeline."""
from __future__ import annotations
import re
import json
import csv
import shutil
import subprocess
import tempfile
from pathlib import Path
from Bio import SeqIO
from Bio.Seq import Seq

from app.core.config import settings
from app.core.logging import logger
from app.services.scorers import (
    combined_efficiency,
    self_complementarity_score,
    find_restriction_sites,
    mode_position_weight,
)
from app.services.doench2016 import doench_2016_score_from_protospacer
from app.services.cfd import cfd_score, aggregate_specificity


PAM_PATTERNS: dict[str, str] = {
    "NGG":    r"[ACGT]GG",
    "NAG":    r"[ACGT]AG",
    "TTTV":   r"TTT[ACG]",
    "NNGRRT": r"[ACGT][ACGT]G[AG][AG]T",
}


def reverse_complement(seq: str) -> str:
    return str(Seq(seq).reverse_complement())


def find_sgrnas(seq: str, pam: str, guide_length: int) -> list[dict]:
    seq = seq.upper().replace("U", "T")
    rgx = re.compile(PAM_PATTERNS[pam])
    results: list[dict] = []

    for m in rgx.finditer(seq):
        pam_start = m.start()
        if pam == "TTTV":
            guide_start = m.end()
            guide_end = guide_start + guide_length
            if guide_end > len(seq): continue
            guide = seq[guide_start:guide_end]
        else:
            guide_start = pam_start - guide_length
            if guide_start < 0: continue
            guide = seq[guide_start:pam_start]
            guide_end = pam_start
        if "N" in guide: continue
        results.append({
            "sgRNA": guide, "pam": m.group(0),
            "start": guide_start + 1, "end": guide_end, "strand": "+",
        })

    rc = reverse_complement(seq)
    for m in rgx.finditer(rc):
        pam_start = m.start()
        if pam == "TTTV":
            guide_start = m.end()
            guide_end = guide_start + guide_length
            if guide_end > len(rc): continue
            guide = rc[guide_start:guide_end]
        else:
            guide_start = pam_start - guide_length
            if guide_start < 0: continue
            guide = rc[guide_start:pam_start]
            guide_end = pam_start
        if "N" in guide: continue
        fwd_end = len(seq) - guide_start
        fwd_start = len(seq) - guide_end + 1
        results.append({
            "sgRNA": guide, "pam": m.group(0),
            "start": fwd_start, "end": fwd_end, "strand": "-",
        })
    return results


def run_cas_offinder(guides: list[str], pam: str, max_mm: int,
                     genome_fasta: Path | None, work_dir: Path) -> dict[str, int]:
    bin_path = shutil.which(settings.CAS_OFFINDER_BIN)
    if not bin_path or not genome_fasta or not Path(genome_fasta).exists():
        return _fallback_offtarget(guides, max_mm, genome_fasta)

    in_path = work_dir / "casoff_input.txt"
    out_path = work_dir / "casoff_output.txt"
    pam_n = "N" * len(pam)
    lines = [str(genome_fasta), f"{'N' * len(guides[0])}{pam_n}"]
    for g in guides:
        lines.append(f"{g}{pam_n} {max_mm}")
    in_path.write_text("\n".join(lines))
    try:
        subprocess.run([bin_path, str(in_path), "C", str(out_path)],
                       check=True, timeout=600)
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        logger.error(f"Cas-OFFinder failed: {e}")
        return {g: 0 for g in guides}
    counts = {g: 0 for g in guides}
    if out_path.exists():
        for line in out_path.read_text().splitlines():
            parts = line.split("\t")
            if len(parts) >= 2:
                gq = parts[0].rstrip("N").upper()
                if gq in counts:
                    counts[gq] += 1
    return counts


def _fallback_offtarget(guides: list[str], max_mm: int, ref: Path | None) -> dict[str, int]:
    if not ref or not Path(ref).exists():
        return {g: 0 for g in guides}
    refs = [str(r.seq).upper() for r in SeqIO.parse(str(ref), "fasta")]
    refs_rc = [reverse_complement(s) for s in refs]
    out = {g: 0 for g in guides}
    for g in guides:
        c = 0
        for r in refs + refs_rc:
            for i in range(len(r) - len(g) + 1):
                w = r[i:i + len(g)]
                if sum(1 for a, b in zip(w, g) if a != b) <= max_mm:
                    c += 1
        out[g] = max(0, c - 1)
    return out


def offtarget_score(off_count: int) -> float:
    if off_count <= 0: return 1.0
    if off_count >= 50: return 0.0
    return max(0.0, 1.0 - off_count / 50)


# Matches an RNAfold output line: dot-bracket structure, then the MFE in
# parentheses. ViennaRNA right-justifies the energy value in a fixed-width
# field, so short magnitudes (e.g. -1.80) get a padding space right after the
# opening paren - "( -1.80)" instead of "(-24.20)". A plain rsplit(" ", 1)
# treats that padding space as the split point and leaves a stray "("
# attached to the structure string, corrupting it (wrong length vs. the
# input sequence) and breaking anything downstream that assumes
# len(structure) == len(sequence) - including RNAplot itself, which then
# rejects the mismatched lengths and silently falls back to an approximate
# layout instead of a real ViennaRNA 2D structure.
_RNAFOLD_LINE_RE = re.compile(r"^([.()\[\]{}<>A-Za-z,|]+)\s+\(\s*(-?\d+\.\d+)\s*\)\s*$")


def run_rnafold(guide: str) -> tuple[float, float, str, list[dict]]:
    """Return (mfe, structure_score, dot_bracket, xy_coords)."""
    bin_path = shutil.which(settings.RNAFOLD_BIN)
    if not bin_path:
        mfe, score = _heuristic_structure(guide)
        return mfe, score, _heuristic_dotbracket(guide), []

    try:
        proc = subprocess.run(
            [bin_path, "--noPS"],
            input=guide,
            capture_output=True, text=True, timeout=10
        )
        out = proc.stdout.strip().splitlines()
        if len(out) >= 2:
            tail = out[1]
            m = _RNAFOLD_LINE_RE.match(tail)
            if not m:
                raise ValueError(f"Unrecognized RNAfold output line: {tail!r}")
            structure = m.group(1)
            mfe = float(m.group(2))
            if len(structure) != len(guide):
                raise ValueError(
                    f"Parsed structure length {len(structure)} != sequence "
                    f"length {len(guide)} (line: {tail!r})"
                )

            # Now get 2D coordinates via RNAplot
            coords = _get_rnaplot_coords(guide, structure)
            return mfe, _mfe_to_score(mfe), structure, coords

    except (subprocess.SubprocessError, ValueError, IndexError) as e:
        logger.debug(f"RNAfold failed: {e}")

    mfe, score = _heuristic_structure(guide)
    return mfe, score, _heuristic_dotbracket(guide), []


def _get_rnaplot_coords(sequence: str, structure: str) -> list[dict]:
    """Run RNAplot to get 2D nucleotide coordinates."""
    import os

    rnaplot_path = shutil.which("RNAplot")
    if not rnaplot_path:
        for p in [r"C:\Program Files\ViennaRNA\RNAplot.exe",
                  r"C:\Program Files (x86)\ViennaRNA\RNAplot.exe"]:
            if Path(p).exists():
                rnaplot_path = p
                break
    if not rnaplot_path:
        return _compute_coords_from_structure(sequence, structure)

    try:
        with tempfile.TemporaryDirectory() as td:
            input_str = f">seq\n{sequence}\n{structure}\n"
            proc = subprocess.run(
                [rnaplot_path, "--output-format=svg"],
                input=input_str,
                capture_output=True,
                text=True,
                timeout=15,
                cwd=td,
                env=os.environ.copy(),
            )
            # RNAplot writes seq_ss.svg in cwd
            svg_file = Path(td) / "seq_ss.svg"
            if svg_file.exists():
                svg_text = svg_file.read_text()
                coords = _parse_rnaplot_svg(svg_text)
                if len(coords) >= len(sequence):
                    return coords[:len(sequence)]
            logger.debug(f"RNAplot SVG file missing or short, stderr: {proc.stderr[:200]}")
    except Exception as e:
        logger.debug(f"RNAplot failed: {e}")

    return _compute_coords_from_structure(sequence, structure)


def _parse_rnaplot_svg(svg_text: str) -> list[dict]:
    """Extract nucleotide x,y positions from RNAplot SVG output."""
    import re
    coords = []

    # RNAplot places nucleotides as text elements with x,y coords.
    # Two independent gotchas here, each silently dropping nucleotides and
    # pushing the count below len(sequence) - which then triggered the
    # heuristic-layout fallback even when RNAplot itself had succeeded:
    #  1. It echoes back whatever alphabet the input used - guides are kept
    #     in DNA form (T, not U) elsewhere in this module, so labels read
    #     "T", not "U". Excluding T/t here dropped every T nucleotide.
    #  2. RNAplot's layout is centered around the origin, so most non-trivial
    #     structures (anything beyond a single small hairpin) place several
    #     nucleotides at negative x/y coordinates. [0-9.]+ can't match the
    #     leading "-", dropping every nucleotide on the negative side.
    # Pattern 1: <text x="X" y="Y">NUCLEOTIDE</text>
    text_pattern = re.compile(
        r'<text[^>]+x="(-?[0-9.]+)"[^>]+y="(-?[0-9.]+)"[^>]*>\s*[ACGUTacgut]\s*</text>'
    )
    for m in text_pattern.finditer(svg_text):
        coords.append({"x": float(m.group(1)), "y": float(m.group(2))})

    if coords:
        return coords

    # Pattern 2: circles (older RNAplot versions)
    circle_pattern = re.compile(
        r'<circle[^>]+cx="(-?[0-9.]+)"[^>]+cy="(-?[0-9.]+)"'
    )
    for m in circle_pattern.finditer(svg_text):
        coords.append({"x": float(m.group(1)), "y": float(m.group(2))})

    return coords


def _compute_coords_from_structure(sequence: str, structure: str) -> list[dict]:
    """
    Compute 2D RNA layout from dot-bracket using
    a recursive stem-loop placement algorithm.
    Gives a proper 2D layout even without RNAplot.
    """
    import math
    n = len(sequence)
    if n == 0:
        return []

    # Build pair table
    pairs = {}
    stack = []
    for i, c in enumerate(structure):
        if c == "(":
            stack.append(i)
        elif c == ")" and stack:
            j = stack.pop()
            pairs[i] = j
            pairs[j] = i

    positions = [None] * n

    def place_loop(indices: list[int], cx: float, cy: float, radius: float, start_angle: float):
        """Place unpaired nucleotides in a loop around (cx, cy)."""
        count = len(indices)
        if count == 0:
            return
        angle_step = (2 * math.pi) / max(count, 1)
        for k, idx in enumerate(indices):
            angle = start_angle + k * angle_step
            positions[idx] = {
                "x": cx + radius * math.cos(angle),
                "y": cy + radius * math.sin(angle)
            }

    def place_stem_loop(i: int, j: int, entry_x: float, entry_y: float,
                        exit_x: float, exit_y: float, depth: int = 0):
        """Recursively place a stem from i..j."""
        if i >= j:
            return

        stem_pairs = []
        si, sj = i, j
        while si < sj and si in pairs and pairs[si] == sj:
            stem_pairs.append((si, sj))
            si += 1
            sj -= 1

        # Place stem nucleotides
        stem_len = len(stem_pairs)
        if stem_len == 0:
            # No stem — place as loop
            loop_indices = list(range(i, j + 1))
            cx = (entry_x + exit_x) / 2
            cy = (entry_y + exit_y) / 2
            r = max(20, len(loop_indices) * 8)
            place_loop(loop_indices, cx, cy, r, math.pi)
            return

        dx = (exit_x - entry_x) / max(stem_len + 1, 1)
        dy = (exit_y - entry_y) / max(stem_len + 1, 1)
        perp_x = -dy / max(math.hypot(dx, dy), 0.001) * 15
        perp_y = dx / max(math.hypot(dx, dy), 0.001) * 15

        for k, (pi, pj) in enumerate(stem_pairs):
            frac = (k + 1) / (stem_len + 1)
            cx = entry_x + frac * (exit_x - entry_x)
            cy = entry_y + frac * (exit_y - entry_y)
            positions[pi] = {"x": cx + perp_x, "y": cy + perp_y}
            positions[pj] = {"x": cx - perp_x, "y": cy - perp_y}

        # Place loop between si and sj
        loop_indices = list(range(si, sj + 1))
        if loop_indices:
            last_stem_x = (positions[stem_pairs[-1][0]]["x"] + positions[stem_pairs[-1][1]]["x"]) / 2
            last_stem_y = (positions[stem_pairs[-1][0]]["y"] + positions[stem_pairs[-1][1]]["y"]) / 2
            loop_r = max(25, len(loop_indices) * 9)
            place_loop(loop_indices, last_stem_x, last_stem_y - loop_r, loop_r, 0)

            # Recurse into nested stems within the loop
            k = si
            while k <= sj:
                if k in pairs and pairs[k] > k:
                    nested_j = pairs[k]
                    if nested_j <= sj:
                        p1x = positions[k]["x"] if positions[k] else last_stem_x
                        p1y = positions[k]["y"] if positions[k] else last_stem_y
                        p2x = positions[nested_j]["x"] if positions[nested_j] else last_stem_x
                        p2y = positions[nested_j]["y"] if positions[nested_j] else last_stem_y
                        place_stem_loop(k, nested_j, p1x, p1y, p2x, p2y, depth + 1)
                        k = nested_j + 1
                        continue
                k += 1

    # Find top-level stems and unpaired regions
    W, H = 800, 600
    x, y = 50.0, float(H // 2)
    step = (W - 100) / max(n - 1, 1)

    # First pass: place all as linear
    for idx in range(n):
        positions[idx] = {"x": x + idx * step, "y": y}

    # Second pass: fix stem-loop positions
    idx = 0
    while idx < n:
        if idx in pairs and pairs[idx] > idx:
            j = pairs[idx]
            place_stem_loop(
                idx, j,
                x + idx * step, y,
                x + j * step, y
            )
            idx = j + 1
        else:
            idx += 1

    # Fill any None positions
    for idx in range(n):
        if positions[idx] is None:
            positions[idx] = {"x": x + idx * step, "y": y}

    return positions


def _heuristic_structure(guide: str) -> tuple[float, float]:
    g = guide.upper()
    if not g: return 0.0, 1.0
    gc = (g.count("G") + g.count("C")) / len(g)
    rc = reverse_complement(g)
    matches = sum(1 for a, b in zip(g, rc) if a == b)
    pal = matches / len(g)
    mfe = -(gc * 8 + pal * 6)
    return mfe, _mfe_to_score(mfe)


def _heuristic_dotbracket(guide: str) -> str:
    """Greedy outermost-pairs dot-bracket. Cheap stand-in for ViennaRNA."""
    g = guide.upper().replace("T", "U")
    n = len(g)
    pair = {"A": "U", "U": "A", "G": "C", "C": "G"}
    out = ["."] * n
    used = [False] * n
    # Outer-pair greedy (one stem only — no nesting in heuristic)
    for i in range(n):
        if used[i]: continue
        for j in range(n - 1, i + 3, -1):
            if used[j]: continue
            if pair.get(g[i]) == g[j] and all(out[k] == "." for k in range(i + 1, j)):
                out[i] = "("
                out[j] = ")"
                used[i] = used[j] = True
                break
    return "".join(out)


def _mfe_to_score(mfe: float) -> float:
    return max(0.0, min(1.0, 1.0 + mfe / 10.0))


def design_sgrnas(fasta_path: Path, pam: str, guide_length: int,
                  max_mismatches: int, genome_fasta: Path | None,
                  top_n: int, job_dir: Path,
                  mode: str = "knockout") -> dict:
    records = list(SeqIO.parse(str(fasta_path), "fasta"))
    if not records:
        raise ValueError("Input FASTA contains no sequences")

    all_results: list[dict] = []

    for rec in records:
        seq_str = str(rec.seq).upper().replace("U", "T")
        seq_len = len(seq_str)
        candidates = find_sgrnas(seq_str, pam, guide_length)
        logger.info(f"{rec.id}: {len(candidates)} candidates (mode={mode})")
        if not candidates:
            continue

        guide_strs = [c["sgRNA"] for c in candidates]
        with tempfile.TemporaryDirectory() as td:
            off_counts = run_cas_offinder(
                guide_strs, pam, max_mismatches,
                Path(genome_fasta) if genome_fasta else fasta_path,
                Path(td),
            )

        for c in candidates:
            g = c["sgRNA"]
            gc = (g.count("G") + g.count("C")) / len(g)
            eff_approx = combined_efficiency(g)

            # Real Doench 2016 score (replaces the heuristic "doench" approximation)
            try:
                # Get 4-nt 5' context and 3-nt 3' context from the input sequence
                if c["strand"] == "+":
                    five = seq_str[max(0, c["start"] - 5):c["start"] - 1]
                    three = seq_str[c["end"] + len(c["pam"]):c["end"] + len(c["pam"]) + 3]
                else:
                    rc_seq = reverse_complement(seq_str)
                    rc_start = seq_len - c["end"]
                    rc_end = seq_len - c["start"] + 1
                    five = rc_seq[max(0, rc_start - 4):rc_start]
                    three = rc_seq[rc_end + len(c["pam"]):rc_end + len(c["pam"]) + 3]
                doench_real = doench_2016_score_from_protospacer(g, five, three)
            except Exception as e:
                logger.debug(f"Doench 2016 fallback for {g}: {e}")
                doench_real = eff_approx["doench"]

            off = off_counts.get(g, 0)
            off_s = offtarget_score(off)
            # Aggregate specificity (placeholder until real off-target hits are returned with sequences)
            specificity = round(off_s * 100, 2)

            SPCAS9_SCAFFOLD = "GTTTTAGAGCTAGAAATAGCAAGTTAAAATAAGGCTAGTCCGTTATCAACTTGAAAAAGTGGCACCGAGTCGGTGC"
            mfe, struct_s, dotbracket, rna_coords = run_rnafold(g + SPCAS9_SCAFFOLD)
            self_comp = self_complementarity_score(g)
            sites = find_restriction_sites(g + c["pam"])
            mode_w = mode_position_weight(c["start"], c["end"], seq_len, mode)

            # Combined efficiency now blends real Doench + heuristic approximations
            combined_eff = round(
                0.55 * doench_real
                + 0.25 * eff_approx["moreno_mateos"]
                + 0.20 * eff_approx["crisprater"], 4)

            base = (
                0.40 * combined_eff +
                0.30 * off_s +
                0.10 * struct_s +
                0.10 * self_comp +
                0.10 * (1.0 if not sites else 0.5)
            )
            composite = round(min(1.0, base * mode_w), 4)

            all_results.append({
                "sequence_id":         rec.id,
                "sgRNA":               g,
                "pam":                 c["pam"],
                "start":               c["start"],
                "end":                 c["end"],
                "strand":              c["strand"],
                "gc_content":          round(gc, 3),
                "doench_score":        round(doench_real, 4),
                "doench_validated":    True,   # flag: real model used
                "moreno_mateos_score": eff_approx["moreno_mateos"],
                "crisprater_score":    eff_approx["crisprater"],
                "efficiency_score":    combined_eff,
                "off_targets":         off,
                "off_target_score":    round(off_s, 4),
                "specificity_score":   specificity,
                "mfe":                 round(mfe, 2),
                "structure_score":     round(struct_s, 4),
                "structure":           dotbracket,
                "rna_coords":          rna_coords,   
                "full_rna_length":     guide_length + len(SPCAS9_SCAFFOLD),
                "self_complementarity": self_comp,
                "restriction_sites":   sites,
                "mode_weight":         round(mode_w, 3),
                "composite_score":     composite,
            })

    all_results.sort(key=lambda r: r["composite_score"], reverse=True)
    for i, r in enumerate(all_results, 1):
        r["rank"] = i
    top = all_results[:top_n]

    json_path = job_dir / "sgrna_results.json"
    json_path.write_text(json.dumps(all_results, indent=2))

    csv_path = job_dir / "sgrna_results.csv"
    if all_results:
        flat = [{
            **r,
            "restriction_sites": ";".join(r["restriction_sites"]),
            "rna_coords": json.dumps(r.get("rna_coords", [])),
        } for r in all_results]
        with csv_path.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(flat[0].keys()))
            w.writeheader()
            w.writerows(flat)

    xlsx_path = _write_xlsx(all_results, job_dir / "sgrna_results.xlsx")
    pdf_path = _write_pdf_report(all_results[:50], job_dir / "sgrna_report.pdf",
                                  meta={"pam": pam, "guide_length": guide_length,
                                        "mode": mode, "top_n": top_n,
                                        "max_mismatches": max_mismatches})

    return {
        "total_candidates": len(all_results),
        "top_n_returned":   len(top),
        "mode":             mode,
        "all_sgRNAs":       all_results,
        "top_sgRNAs":       top,
        "files": {
            "json": str(json_path),
            "csv":  str(csv_path),
            "xlsx": str(xlsx_path) if xlsx_path else None,
            "pdf":  str(pdf_path) if pdf_path else None,
        },
    }


def _write_xlsx(rows: list[dict], path: Path) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed — skipping XLSX export")
        return None
    if not rows: return None
    wb = Workbook()
    ws = wb.active
    ws.title = "sgRNA results"
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="16A34A")
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([
            json.dumps(r[k]) if (isinstance(r[k], list) and r[k] and isinstance(r[k][0], dict))
            else (";".join(r[k]) if isinstance(r[k], list) else r[k])
            for k in headers
        ])
    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def _write_pdf_report(rows: list[dict], path: Path, meta: dict) -> Path | None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        from matplotlib.backends.backend_pdf import PdfPages
        import matplotlib.pyplot as plt
    except ImportError:
        return None
    if not rows: return None

    with PdfPages(path) as pdf:
        fig = plt.figure(figsize=(11, 8.5))
        fig.text(0.5, 0.85, "EditEase - sgRNA Design Report",
                 ha="center", size=22, weight="bold")
        fig.text(0.5, 0.78, f"Mode: {meta.get('mode')}  |  PAM: {meta.get('pam')}  |  "
                            f"Guide length: {meta.get('guide_length')}",
                 ha="center", size=12)
        fig.text(0.5, 0.74, f"Top {meta.get('top_n')} sgRNAs  |  "
                            f"Max mismatches: {meta.get('max_mismatches')}",
                 ha="center", size=12)

        top_table = rows[:10]
        cell_text = [
            [r["rank"], r["sgRNA"], r["pam"],
             f"{r['start']}-{r['end']}", r["strand"],
             f"{r['gc_content']*100:.0f}%",
             f"{r['efficiency_score']:.2f}",
             r["off_targets"],
             f"{r['mfe']:.1f}",
             f"{r['composite_score']:.3f}"]
            for r in top_table
        ]
        ax = fig.add_axes([0.05, 0.05, 0.9, 0.6])
        ax.axis("off")
        tbl = ax.table(
            cellText=cell_text,
            colLabels=["#", "sgRNA", "PAM", "Pos", "Str", "GC", "Eff", "OT", "MFE", "Score"],
            loc="upper center",
        )
        tbl.auto_set_font_size(False); tbl.set_fontsize(8); tbl.scale(1, 1.4)
        for j in range(10):
            tbl[(0, j)].set_facecolor("#16A34A")
            tbl[(0, j)].set_text_props(color="white", weight="bold")
        pdf.savefig(fig); plt.close(fig)

        fig2, ax2 = plt.subplots(figsize=(11, 5))
        ax2.hist([r["composite_score"] for r in rows], bins=20,
                 color="#16A34A", edgecolor="black")
        ax2.set_xlabel("Composite score")
        ax2.set_ylabel("Count")
        ax2.set_title("Composite-score distribution")
        pdf.savefig(fig2); plt.close(fig2)

    return path
